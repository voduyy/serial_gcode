import insightface
import cv2
import numpy as np
import warnings
warnings.filterwarnings("ignore", category=FutureWarning)
import rembg
from PIL import Image
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from openpyxl.styles import Font
import matplotlib.pyplot as plt
from sklearn.cluster import DBSCAN
import sys
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '.')))
import plot_gcode
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import global_var
import warnings
warnings.filterwarnings("ignore", category=UserWarning)
# Adaptive Resampling (góc + khoảng cách)
def simplify_and_adaptive_resample(points, simplify_epsilon=1.0, angle_thresh=10, min_spacing=4):
    if len(points) < 3:
        return points
    # Douglas-Peucker simplification cv2.approxPolyDP
    approx = cv2.approxPolyDP(points, epsilon=simplify_epsilon, closed=True)
    if len(approx) < 3:
        approx = points

    approx = approx.squeeze()
    keep_points = [approx[0]]

    for i in range(1, len(approx) - 1):
        p_prev = approx[i - 1]
        p_curr = approx[i]
        p_next = approx[i + 1]

        v1 = p_curr - p_prev
        v2 = p_next - p_curr

        cos_angle = np.dot(v1, v2) / (np.linalg.norm(v1) * np.linalg.norm(v2) + 1e-8)
        angle = np.degrees(np.arccos(np.clip(cos_angle, -1.0, 1.0)))

        dist = np.linalg.norm(p_curr - keep_points[-1])
        if angle < (180 - angle_thresh) or dist >= min_spacing:
            keep_points.append(p_curr)

    keep_points.append(approx[-1])
    return np.array(keep_points, dtype=np.int32).reshape(-1, 1, 2)

# --- Class xử lý ảnh và sinh G-code ---
class Picture:
    def __init__(self, filepath, x_max=100, y_max=100):
        self.img = Image.open(filepath).convert("RGB")
        self.img = np.array(self.img)
        self.h, self.w, self.c = self.img.shape
        self.pre = np.ones(self.img.shape)
        self.gcode = ['G28']
        self.x_max = x_max
        self.y_max = y_max

    def gray_scale(self):
        gray = cv2.cvtColor(self.img, cv2.COLOR_RGB2GRAY)
        self.gray = gray / 255.0  # For visualization

        blurred = cv2.GaussianBlur(gray, (3, 3), 0)
        edges = cv2.Canny(blurred, threshold1=50, threshold2=150)

        binary = (edges / 255.0).astype(float)
        # binary = 1.0 - (edges / 255.0).astype(float)  # <- đảo ngược để có nền trắng
        self.pre = np.stack([binary] * 3, axis=-1)
        return self.pre

    def save_gray(self, output):
        if hasattr(self, 'gray') and self.gray is not None:
            plt.imshow(self.gray, cmap='gray')
            plt.axis('off')
            plt.imsave(output + '_gray.jpg', self.gray, cmap='gray')
            # plt.imsave(os.path.normpath(output + '_gray.jpg'), self.gray, cmap='gray')
            print('✅ Saved ' + output + '_gray.jpg')

    def save_binary(self, output):

        if hasattr(self, 'pre') and self.pre is not None:
            # Đảo màu: nét đen (0), nền trắng (255)
            binary_inverted = (1 - self.pre)  # đảo ngược: trắng → đen, đen → trắng
            plt.imshow(binary_inverted, cmap='gray')
            plt.axis('off')
            plt.imsave(output + '_binary.jpg', binary_inverted, cmap='gray')
            print('✅ Saved ' + output + '_binary.jpg ')

    def gen_gcode(self, eps=10, simplify_epsilon=1, min_spacing=4, min_contour_len=10):
        binary = (self.pre[:, :, 0] > 0.5).astype(np.uint8) * 255
        contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
        ratio = self.x_max / max(self.w, self.h)
        total_points = 0

        centers = []
        valid_contours = []
        for contour in contours:
            if len(contour) < min_contour_len:
                continue
            M = cv2.moments(contour)
            if M["m00"] != 0:
                cx = int(M["m10"] / M["m00"])
                cy = int(M["m01"] / M["m00"])
            else:
                cx, cy = contour[0][0]
            centers.append([cx, cy])
            valid_contours.append(contour)

        if not centers:
            return self.gcode, total_points

        labels = DBSCAN(eps=eps, min_samples=1).fit_predict(centers)

        clusters = {}
        for label, contour in zip(labels, valid_contours):
            clusters.setdefault(label, []).append(contour)

        spindle_on = False  # trạng thái spindle

        for cluster in clusters.values():
            for contour in cluster:
                simplified = simplify_and_adaptive_resample(
                    contour,
                    simplify_epsilon=simplify_epsilon,
                    angle_thresh=15,
                    min_spacing=min_spacing
                )
                if len(simplified) < 2:
                    continue
                total_points += len(simplified)

                # Di chuyển nhanh đến điểm đầu (G0 + M5)
                x0, y0 = simplified[0][0]
                y0_flipped = self.h - y0
                if spindle_on:
                    self.gcode.append("M5")  # Tắt spindle trước di chuyển nhanh
                    spindle_on = False
                x0_out, y0_out = x0 * ratio, y0_flipped * ratio
                self.gcode.append(f"G0 X{x0_out:.2f} Y{y0_out:.2f}")
                self.gcode.append(f"G92 X{x0_out:.2f} Y{y0_out:.2f}")
                 # Vẽ (G1 + M3)
                for pt in simplified[1:]:
                    x, y = pt[0]
                    y_flipped = self.h - y
                    if not spindle_on:
                        self.gcode.append("M3")  # Bật spindle trước khi vẽ
                        spindle_on = True
                    x_out, y_out = x * ratio, y_flipped * ratio
                    self.gcode.append(f"G1 X{x_out:.2f} Y{y_out:.2f}")
                    self.gcode.append(f"G92 X{x_out:.2f} Y{y_out:.2f}")

        # Kết thúc nếu spindle đang bật thì tắt đi
        if spindle_on:
            self.gcode.append("M5")
            spindle_on = False

        return self.gcode, total_points

    def save_gcode(self, output_name):
        os.makedirs(os.path.dirname(output_name), exist_ok=True)
        with open(f'{output_name}_gcode.nc', 'w') as f:
            for line in self.gcode:
                f.write(f'{line}\n')
        print(f'✅ Saved {output_name}_gcode.nc')

# --- Tiện ích resize ảnh trước khi gắn vào Excel ---
def resize_and_save_temp(image_path, output_path, max_size=(100, 100)):
    try:
        img = PILImage.open(image_path).convert("RGB")  # ← Thêm convert tại đây
        img.thumbnail(max_size)
        img.save(output_path)
        return True
    except Exception as e:
        print(f"⚠️ Lỗi khi resize ảnh: {image_path} → {e}")
        return False

# --- Xử lý khuôn mặt ---
def init_face_analyzer():
    # print("Initializing face analysis...")
    face_analyzer = insightface.app.FaceAnalysis(name='buffalo_l')
    # try:
    #     face_analyzer.prepare(ctx_id=0)
    #     print(f"✅ Using GPU: {face_analyzer.models['detection'].providers}")
    # except Exception:
    #     print("GPU initialization failed. Switching to CPU...")
    #
    face_analyzer.prepare(ctx_id=-1)
    return face_analyzer

def align_face(image, face):
    left_eye = tuple(face.kps[0].astype(int))
    right_eye = tuple(face.kps[1].astype(int))
    eyes_center = tuple(map(int, ((left_eye[0] + right_eye[0]) / 2, (left_eye[1] + right_eye[1]) / 2)))
    dx = right_eye[0] - left_eye[0]
    dy = right_eye[1] - left_eye[1]
    angle = np.degrees(np.arctan2(dy, dx))
    rotation_matrix = cv2.getRotationMatrix2D(eyes_center, angle, 1)
    aligned = cv2.warpAffine(image.astype(np.float32), rotation_matrix, (image.shape[1], image.shape[0]),
                             flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_CONSTANT, borderValue=(255, 255, 255))
    return np.clip(aligned, 0, 255).astype(np.uint8)

def resize_to_a4(image, target_width=300, target_height=300):
    h, w = image.shape[:2]
    scale = min(target_width / w, target_height / h)
    new_w, new_h = int(w * scale), int(h * scale)
    resized = cv2.resize(image, (new_w, new_h), interpolation=cv2.INTER_AREA)
    canvas = np.ones((target_height, target_width, 3), dtype=np.uint8) * 255
    top = (target_height - new_h) // 2
    left = (target_width - new_w) // 2
    canvas[top:top + new_h, left:left + new_w] = resized
    return canvas

# === Main pipeline ===
def main():
    face_model = init_face_analyzer()

    input_face_dir = "input_image"
    filenames = ""
    # mode_manual = '1'
    # if not global_var.is_capture and not global_var.is_choose_image:
    #     if mode_manual == '1': #handle capture
    #         path = os.path.join(input_face_dir,"capture")
    #         filenames = sorted(
    #             [f for f in os.listdir(path) if f.lower().endswith(('.jpg', '.jpeg', '.png'))],
    #             key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else float('inf')
    #         )
    #     elif mode_manual == '2': #handle test
    #         path = os.path.join(input_face_dir, "test")
    #         filenames = sorted(
    #             [f for f in os.listdir(path) if f.lower().endswith(('.jpg', '.jpeg', '.png'))],
    #             key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else float('inf')
    #         )
    #     elif mode_manual == '3':  # handle test
    #         path = os.path.join(input_face_dir, "shape")
    #         filenames = sorted(
    #             [f for f in os.listdir(path) if f.lower().endswith(('.jpg', '.jpeg', '.png'))],
    #             key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else float('inf')
    #         )
    #     else:
    #         print("Chế độ không hợp lệ.")
    #         exit()
         # handle which being capture or select
    path = os.path.join(input_face_dir, "capture")
    filenames = global_var.image_name

    output_image_folder = 'Image2Gcode\output_image'
    output_gcode_folder = 'Image2Gcode\output_gcode'
    image_grayscale_folder="Image2Gcode\image_grayscale"
    excel_folder = 'Image2Gcode\excel'
    thumb_folder = os.path.join(excel_folder, 'thumbs')
    img_crop_folder = "Image2Gcode\image_crop"

    os.makedirs(output_image_folder, exist_ok=True)
    os.makedirs(thumb_folder, exist_ok=True)
    os.makedirs(excel_folder, exist_ok=True)
    os.makedirs(img_crop_folder, exist_ok=True)
    os.makedirs(output_gcode_folder, exist_ok=True)
    os.makedirs(image_grayscale_folder, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Thời gian xử lý ảnh"
    ws.append(["Tên file", "Input", "Thời gian (s)", "Output", "Số điểm G-code"])
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # filenames = sorted(
    #     [f for f in os.listdir(input_face_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png'))],
    #     key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else float('inf')
    # )
    if not isinstance(filenames, list):
        filenames = [filenames]

    for file in filenames:
        print(f"\n📂 Đang xử lý ảnh: {file}")
        start_time = datetime.now()
        base_dir = os.path.dirname(os.path.abspath(__file__))
        input_path = os.path.join(base_dir, path, file)
        image = cv2.imread(input_path)
        if image is None:
            print(f"⚠️ Không đọc được ảnh: {file}")
            continue

        # Tách nền
        image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        pil_img = Image.fromarray(image_rgb)
        removed = rembg.remove(pil_img)
        rgba = removed.convert("RGBA")
        np_img = np.array(rgba)
        white_bg = np.ones_like(np_img[:, :, :3], dtype=np.uint8) * 255
        alpha = np_img[:, :, 3:4] / 255.0
        blend = (np_img[:, :, :3] * alpha + white_bg * (1 - alpha)).astype(np.uint8)

        # Phát hiện mặt
        faces = face_model.get(blend)
        if not faces:
            print(f"❌ Không phát hiện khuôn mặt trong ảnh: {file}")
            continue

        face = faces[0]
        aligned = align_face(blend, face)

        # === CHÈN ĐOẠN NÀY VÀO ĐÂY ===
        bounding_box_dir = "Image2Gcode\image_bounding_box"
        os.makedirs(bounding_box_dir, exist_ok=True)

        boxed_img = blend.copy()
        x1, y1, x2, y2 = face.bbox.astype(int)
        cv2.rectangle(boxed_img, (x1, y1), (x2, y2), (0, 255, 0), 2)
        for (x, y) in face.kps.astype(int):
            cv2.circle(boxed_img, (x, y), 2, (0, 0, 255), -1)
        file_name = os.path.splitext(file)[0]
        bbox_save_path = os.path.join(bounding_box_dir, file_name + "_bbox.jpg")
        cv2.imwrite(bbox_save_path, cv2.cvtColor(boxed_img, cv2.COLOR_RGB2BGR))
        print(f"🟩 Bounding box image saved: {os.path.basename(bbox_save_path)}")
        # === HẾT ĐOẠN THÊM ===

        x1, y1, x2, y2 = face.bbox.astype(int)
        h, w = aligned.shape[:2]
        top = np.clip(y1 - int(0.45 * (y2 - y1)), 0, h)
        bottom = np.clip(y2 + int(0.05 * (y2 - y1)), 0, h)
        left = np.clip(x1 - int(0.3 * (x2 - x1)), 0, w)
        right = np.clip(x2 + int(0.3 * (x2 - x1)), 0, w)
        crop = aligned[top:bottom, left:right]
        crop_rgb = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)

        # Resize A4
        a4 = resize_to_a4(crop_rgb)
        a4_name = file_name + "_a4.jpg"
        a4_path = os.path.join(img_crop_folder, a4_name)
        cv2.imwrite(a4_path, a4)
        print(f"✅ Aligned and cropped face: {a4_name}")

        # Xử lý nhị phân & G-code
        output_image_name = os.path.join(output_image_folder, file_name)
        image_grayscale_name = os.path.join(image_grayscale_folder, file_name)
        pic = Picture(a4_path)
        pic.gray_scale()  # ← Truyền face vào đây
        # pic.gray_scale()
        pic.save_gray(image_grayscale_name)
        pic.save_binary(output_image_name)

        output_gcode_name = os.path.join(output_gcode_folder, file_name)

        gcode, num_points = pic.gen_gcode()
        pic.save_gcode(output_gcode_name)

        duration = (datetime.now() - start_time).total_seconds()
        ws.append([file, "", round(duration, 3), "", num_points])
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = 85

        thumb1 = os.path.join(thumb_folder, f'{file_name}_thumb1.jpg')
        thumb2 = os.path.join(thumb_folder, f'{file_name}_thumb2.jpg')

        if resize_and_save_temp(input_path, thumb1):
            ws.add_image(XLImage(thumb1), f'B{current_row}')
        if resize_and_save_temp(f'{output_image_name}_binary.jpg', thumb2):
            ws.add_image(XLImage(thumb2), f'D{current_row}')

    total_time = sum(ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1))
    ws.append(["", "", f"TỔNG: {total_time:.3f} giây", "", ""])
    ws[f"C{ws.max_row}"].font = Font(bold=True)

    wb.save(os.path.join(excel_folder, 'time_processing.xlsx'))
    print("\n🎉 Xử lý hoàn tất toàn bộ ảnh!")
    global_var.is_finish_covert_image = True
    image_gcode_name = os.path.splitext(global_var.image_name)[0]
    plot_gcode.draw_gcode_and_save(f"Image2Gcode/output_gcode/{image_gcode_name}_gcode.nc",
                        f"Image2Gcode/simulate_image/{image_gcode_name}.png", angle_deg=180)

if __name__ == '__main__':
    main()