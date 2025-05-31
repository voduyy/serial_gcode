import cv2
import numpy as np
import matplotlib.pyplot as plt
from PIL import Image
import os
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from openpyxl.styles import Font
import mediapipe as mp

# --- Lớp xử lý ảnh và sinh G-code từ ảnh ---
class Picture:
    # Mở ảnh và chuyển sang RGB, lưu các thuộc tính ảnh gốc
    def __init__(self, filepath, x_max=40, y_max=40):
        self.img = Image.open(filepath).convert("RGB")
        self.img = np.array(self.img)
        self.h, self.w, self.c = self.img.shape
        self.pre = np.ones(self.img.shape)
        self.gcode = ['G28']
        self.x_max = x_max
        self.y_max = y_max

    # Chuyển sang grayscale, làm mịn, phát hiện biên bằng Canny
    def gray_scale(self):
        gray = cv2.cvtColor(self.img, cv2.COLOR_RGB2GRAY)
        self.gray = gray / 255.0  # For visualization

        blurred = cv2.GaussianBlur(gray, (3, 3), 0)
        edges = cv2.Canny(blurred, threshold1=50, threshold2=150)

        binary = (edges / 255.0).astype(float)
        # binary = 1.0 - (edges / 255.0).astype(float)  # <- đảo ngược để có nền trắng
        self.pre = np.stack([binary] * 3, axis=-1)
        return self.pre

    # def detect_and_highlight_face(self):
    #     gray_img = cv2.cvtColor(self.img, cv2.COLOR_RGB2GRAY)
    #     face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
    #     faces = face_cascade.detectMultiScale(gray_img, scaleFactor=1.1, minNeighbors=5)
    #
    #     if len(faces) == 0:
    #         print("⚠️ Không tìm thấy khuôn mặt.")
    #         return
    #
    #     (x, y, w, h) = faces[0]
    #     pad_x = int(w * 0.3)
    #     pad_y = int(h * 0.4)
    #     x1, y1 = max(0, x - pad_x), max(0, y - pad_y)
    #     x2, y2 = min(self.w, x + w + pad_x), min(self.h, y + h + pad_y)
    #
    #     mask = np.zeros((self.h, self.w), dtype=np.uint8)
    #     mask[y1:y2, x1:x2] = 255
    #
    #     edges = (self.pre[:, :, 0] * 255).astype(np.uint8)
    #     highlighted = cv2.bitwise_and(edges, edges, mask=mask)
    #
    #     # Làm mịn nhẹ thay vì làm đậm
    #     highlighted = cv2.GaussianBlur(highlighted, (3, 3), sigmaX=0.5)
    #
    #     normalized = (highlighted / 255.0).astype(float)
    #     self.pre = np.stack([normalized] * 3, axis=-1)
    #
    # def detect_details(self):
    #     gray_img = cv2.cvtColor(self.img, cv2.COLOR_RGB2GRAY)
    #     details = []
    #
    #     for name, xml in [
    #         ("eye", 'haarcascade_eye.xml'),
    #         ("nose", 'haarcascade_mcs_nose.xml'),
    #         ("mouth", 'haarcascade_smile.xml')
    #     ]:
    #         xml_path = os.path.join("cascades", xml)
    #         detector = cv2.CascadeClassifier(xml_path)
    #
    #         if detector.empty():
    #             print(f"⚠️ Không thể load '{xml}' — bỏ qua {name}.")
    #             continue
    #
    #         parts = detector.detectMultiScale(gray_img, scaleFactor=1.1, minNeighbors=5)
    #         for (x, y, w, h) in parts:
    #             details.append((x, y, x + w, y + h))
    #
    #     # Tô đậm các vùng phát hiện được
    #     mask = np.zeros((self.h, self.w), dtype=np.uint8)
    #     for (x1, y1, x2, y2) in details:
    #         mask[y1:y2, x1:x2] = 255
    #
    #     edges = (self.pre[:, :, 0] * 255).astype(np.uint8)
    #     highlighted = cv2.bitwise_and(edges, edges, mask=mask)
    #     highlighted = cv2.GaussianBlur(highlighted, (3, 3), sigmaX=0.5)
    #
    #     normalized = (highlighted / 255.0).astype(float)
    #     self.pre = np.maximum(self.pre, np.stack([normalized] * 3, axis=-1))

    # Lưu ảnh grayscale về file JPG
    def save_gray(self, output):
        if hasattr(self, 'gray') and self.gray is not None:
            plt.imshow(self.gray, cmap='gray')
            plt.axis('off')
            plt.imsave(output + '_gray.jpg', self.gray, cmap='gray')
            print('✅ Saved ' + output + '_gray.jpg')

     # Lưu ảnh nhị phân (sketch sau Canny) về file JPG
    def save_binary(self, output):
        plt.imshow(self.pre, cmap='gray')
        plt.axis('off')
        plt.imsave(output + '_binary.jpg', self.pre)
        print('✅ Saved ' + output + '_binary.jpg')

    # Chuyển ảnh nhị phân thành tập lệnh G-code
    def gen_gcode(self):
        binary = (self.pre[:, :, 0] > 0.5).astype(np.uint8) * 255 # -> đen
        # binary = (self.pre[:, :, 0] < 0.5).astype(np.uint8) * 255
        contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
        ratio = self.x_max / max(self.w, self.h)
        total_points = 0  #  biến đếm số điểm G-code

        for contour in contours:
            if len(contour) < 2:
                continue
            total_points += len(contour)
            x0, y0 = contour[0][0]
            y0_flipped = self.h - y0
            # self.gcode.append("M280 P0 S60")
            self.gcode.append(f"G0 X{x0 * ratio:.4f} Y{y0_flipped * ratio:.4f}")
            # self.gcode.append("M280 P0 S0")
            for pt in contour[1:]:
                x, y = pt[0]
                y_flipped = self.h - y
                self.gcode.append(f"G1 X{x * ratio:.4f} Y{y_flipped * ratio:.4f}")
            # self.gcode.append("M280 P0 S60")
        return self.gcode, total_points

    # Ghi tập lệnh G-code ra file .nc
    def save_gcode(self, output_name):
        os.makedirs(os.path.dirname(output_name), exist_ok=True)
        with open(f'{output_name}_gcode.nc', 'w') as f:
            for line in self.gcode:
                f.write(f'{line}\n')
        print(f'✅ Saved {output_name}_gcode.nc')

# Resize ảnh nhỏ lại trước khi chèn vào Excel
def resize_and_save_temp(image_path, output_path, max_size=(100, 100)):
    try:
        img = PILImage.open(image_path)
        img.thumbnail(max_size)
        img.save(output_path)
        return True
    except Exception as e:
        print(f"⚠️ Lỗi khi resize ảnh: {image_path} → {e}")
        return False

if __name__ == '__main__':
    input_folder = 'img'
    output_folder = 'out'
    excel_folder = 'excel'
    thumb_folder = os.path.join(excel_folder, 'thumbs')
    os.makedirs(output_folder, exist_ok=True)
    os.makedirs(thumb_folder, exist_ok=True)
    os.makedirs(excel_folder, exist_ok=True)

    valid_exts = ['.jpg', '.jpeg', '.png', '.bmp']

    def extract_number(filename):
        match = re.search(r'\d+', filename)
        return int(match.group()) if match else float('inf')

    files = sorted(
        [f for f in os.listdir(input_folder) if os.path.splitext(f)[1].lower() in valid_exts],
        key=extract_number
    )

    # Tạo workbook Excel để ghi log xử lý
    wb = Workbook()
    ws = wb.active
    ws.title = "Thời gian xử lý ảnh"
    ws.append(["Tên file", "Input", "Thời gian (s)", "Output", "Số điểm G-code"])
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    print(f"\n🚀 Bắt đầu xử lý toàn bộ...")

    # Vòng lặp xử lý từng ảnh trong thư mục input
    for idx, filename in enumerate(files):
        input_path = os.path.join(input_folder, filename)
        base_name = os.path.splitext(filename)[0]
        output_name = os.path.join(output_folder, f"{idx:02d}_{base_name}")

        print(f"\n🔧 Đang xử lý ảnh: {filename}")
        start_time = datetime.now()

        # Xử lý ảnh và sinh G-code
        pic = Picture(input_path)
        pic.gray_scale()
        pic.save_gray(output_name)
        pic.save_binary(output_name)
        gcode, num_points = pic.gen_gcode()
        pic.save_gcode(output_name)

        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        print(f"✅ Xong ảnh: {filename}")

        ws.append([filename, "", round(duration, 3), "", num_points])
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = 75

        # Tạo và chèn thumbnail ảnh gốc + binary vào Excel
        img_origin_path = input_path
        img_binary_path = f'{output_name}_binary.jpg'
        tmp_img1 = os.path.join(thumb_folder, f'{idx:02d}_{base_name}_thumb1.jpg')
        tmp_img2 = os.path.join(thumb_folder, f'{idx:02d}_{base_name}_thumb2.jpg')

        if os.path.exists(img_origin_path) and resize_and_save_temp(img_origin_path, tmp_img1):
            img1 = XLImage(tmp_img1)
            ws.add_image(img1, f'B{current_row}')

        if os.path.exists(img_binary_path) and resize_and_save_temp(img_binary_path, tmp_img2):
            img2 = XLImage(tmp_img2)
            ws.add_image(img2, f'D{current_row}')

    # Tổng thời gian xử lý
    total_duration = sum([ws.cell(row=i, column=3).value for i in range(2, ws.max_row + 1)])
    print(f"📊 Tổng thời gian xử lý: {total_duration:.3f} giây")

    # Ghi dòng tổng vào cuối cột "Thời gian (s)"
    ws.append(["", "", f"TỔNG: {total_duration:.3f} giây", "", ""])
    summary_row = ws.max_row
    ws[f"C{summary_row}"].font = Font(bold=True)

    # Lưu file Excel tổng kết
    excel_path = os.path.join(excel_folder, 'time_processing.xlsx')
    wb.save(excel_path)
    print(f"📄 Đã lưu dữ liệu xử lý vào: {excel_path}")